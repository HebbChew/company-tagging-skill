#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""终稿合并：Pass A merged + Pass B + 搜索结果 → 终稿Excel + 质检统计。零token。"""
import json, csv, glob, re, collections, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import os
BASE = os.environ.get('MHB_TAG_BASE', os.path.dirname(os.path.abspath(__file__)))
TODAY = '2026-08-27'

# 一级映射
m12, sys2l1 = {}, {}
with open(f'{BASE}/技术标签词表.csv', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        m12[(r['体系'], r['二级'])] = r['一级']
        sys2l1.setdefault(r['体系'], r['一级'])
SYNBIO = {'材料应用', '食品应用', '农牧应用', '其他应用'}
RENAME_L1 = {'生物制造与递送': '合成生物与生物制造', '合成生物学': '合成生物与生物制造'}
# 标签归一映射（method/标签归一.csv）：(原体系,原二级) → (新体系,新一级,新二级,新三级)
TAG_MAP = {}
TAG_MAP3 = {}
with open(f'{BASE}/method/标签归一.csv', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        key2 = r['原二级']
        if '|三级:' in key2:
            base2, l3v = key2.split('|三级:')
            TAG_MAP3[(r['原体系'], base2, l3v)] = (r['新体系'], r['新一级'], r['新二级'], (r['新三级'] or '').strip())
        else:
            TAG_MAP[(r['原体系'], key2)] = (r['新体系'], r['新一级'], r['新二级'], (r['新三级'] or '').strip())
def tag_norm(sys_, l2, l3):
    m3 = TAG_MAP3.get((sys_, l2, l3 or ''))
    if m3: return m3[0], m3[1], m3[2], m3[3], m3[1]
    m = TAG_MAP.get((sys_, l2))
    if m: return m[0], m[1], m[2], (m[3] or l3), m[1]
    return sys_, None, l2, l3, None
SUP_L1 = {'科研试剂耗材': '生物技术与上游试剂耗材设备', '科研仪器设备': '生物技术与上游试剂耗材设备',
          '模型系统与平台': '生物技术与上游试剂耗材设备', '类器官与器官芯片': '生物技术与上游试剂耗材设备'}
def lvl1(s, l2):
    if s == '周边配套': return '周边配套'
    if l2 in ('医药供应链上游配套', '医药流通与供应链'): return RENAME_L1.get('医药供应链与流通（补充）', '医药供应链与流通（补充）')
    if l2 in SYNBIO: return RENAME_L1.get('合成生物学', '合成生物学')
    if l2 in SUP_L1: return SUP_L1[l2]
    v = m12.get((s, l2), sys2l1.get(s, ''))
    return RENAME_L1.get(v, v)

# 人工复核修正表（质检发现的问题逐条修正，优先级最高）
OVERRIDE = {}
import csv as _csv
with open(f'{BASE}/method/人工复核修正.csv', encoding='utf-8-sig') as _f:
    for _r in _csv.DictReader(_f):
        OVERRIDE[(int(_r['序号']), (_r.get('公司名称') or '').strip())] = _r

# 搜索结果
import re as _re
_OURS = _re.compile(r'^(full_S2_part\d+|full_S1_part\d+|full_S0A_A\d+|full_S0B_B\d+|full_S0C_C\d+-\d+|full_S0C_C3a|refind_R\d+|audit_fix|reverify_3家|full_S0A_组|full_S0B_组|full_S0C_批|full_S1_组|full_S2_组|full_S0W_W)')
def _our_search_files():
    import glob as _g
    return [p for p in _g.glob(f'{BASE}/cache/search/*.jsonl') if _OURS.match(p.split('/')[-1])]

S = {}
for p in _our_search_files():
    for l in open(p, encoding='utf-8'):
        try: r = json.loads(l)
        except: continue
        S[r['序号']] = r
# 复找结果后加载，覆盖同序号的空记录（复找=更深搜索，结论优先）
for p in glob.glob(f'{BASE}/cache/search/refind_*.jsonl') + glob.glob(f'{BASE}/cache/search/audit_*.jsonl'):
    for l in open(p, encoding='utf-8'):
        try: r = json.loads(l)
        except: continue
        S[r['序号']] = r

# 兜底句（区分"搜过没有"与"没搜"）
SPECIFIC = re.compile(r'批发|零售|制造|诊所|门诊|医院|检验')
def fallback_intro(hy, searched):
    if not searched:
        return '（注册资本≤100万档，未搜索）'
    if not hy:
        return '（登记行业信息笼统）经搜索暂无公开经营信息'
    if SPECIFIC.search(hy):
        return f'（登记行业：{hy}）经搜索暂无公开经营信息'
    return '（登记行业信息笼统）经搜索暂无公开经营信息'

# 格式归一
def norm_intro(s):
    if not s: return ''
    s = re.sub(r'^(该公司|该企业|这家公司|该企业是|这是一家)', '', s).strip()
    s = re.sub(r'^(是|为)', '', s).strip()
    if s and not s.endswith(('。', '！', '）')): s += '。'
    return s

# 母公司名归一（v2：剥离括号注记为标准名，注记并入备注）
def norm_parent(p):
    if not p: return '', ''
    p = re.sub(r'\s+', '', p).strip()
    m = re.match(r'^(.+?)[（(](.+?)[）)]$', p)
    if m:
        base, note = m.group(1), m.group(2)
        # 括号里是持股/代码/说明类注记 → 剥离；括号是公司名一部分（如"科赴（Kenvue）"）则保留
        if re.search(r'持股|全资|孙公司|子公司|代码|\d{6}|HK|NYSE|收购|预计|待', note):
            return base, f'母公司注记:{note}'
    return p, ''

def canon_parent(p):
    """归一键：去括号+去组织形式后缀，用于同集团聚类"""
    c = re.sub(r'[（(].*?[）)]', '', p)
    c = re.sub(r'(集团股份有限公司|股份有限公司|有限责任公司|有限公司|集团)$', '', c)
    return c

# 地位词弱信源检测
SUPER = re.compile(r'首个|首家|龙头|领先|唯一|最大|首创|头部|领军')
WEAK_SRC = re.compile(r'猎聘|BOSS|顺企|黄页|搜狐|网易|腾讯|公众号|自媒体')
STRONG_SRC = re.compile(r'官网|公告|招股|年报|新华社|新华|人民|政府|巨潮|交易所|CDE|NMPA|企查查|天眼查|爱企查|启信宝')

wb = openpyxl.Workbook(); ws = wb.active; ws.title = '终稿'
HDR = ['序号', '公司名称', '相关度', '边界理由', '主体系', '一级', '二级标签主', '二级标签副', '三级标签',
       '标签置信度', '标签依据', '一句话简介', '行业地位', '母公司', '母公司置信度', '知名度',
       '搜索层级', '信息状态', '证据摘要', '备注', '核验日期']
ws.append(HDR)
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1F6FB2')
    c.alignment = Alignment(horizontal='center', vertical='center')

qc = collections.Counter()
boundary, noinfo_list, doubts = [], [], []
n = 0
gate_dropped = 0
# 母公司别名预扫描：同归一键聚类，映射到该簇最常出现的完整名
_pc = collections.defaultdict(collections.Counter)
for tsv in glob.glob(f'{BASE}/full/batches/batch_f*.tsv'):
    b = tsv.split('/')[-1].replace('.tsv', '')
    for l in open(f'{BASE}/cache/passA_intro/{b}_v3merged.jsonl', encoding='utf-8'):
        a = json.loads(l)
        s = S.get(a['序号'])
        raw = ((s.get('母公司_终稿') if s else '') or a.get('母公司', ''))
        pn, _ = norm_parent(raw)
        if pn and pn != '无':
            _pc[canon_parent(pn)][pn] += 1
PARENT_ALIAS = {}
for ck, cnt in _pc.items():
    best = cnt.most_common(1)[0][0]
    for variant in cnt:
        PARENT_ALIAS[variant] = best
print(f'母公司归一: {len(_pc)} 簇, {sum(len(c) for c in _pc.values())} 个写法')

for tsv in sorted(glob.glob(f'{BASE}/full/batches/batch_f*.tsv')):
    b = tsv.split('/')[-1].replace('.tsv', '')
    with open(tsv, encoding='utf-8') as f:
        header = f.readline().rstrip('\n').split('\t')
        base_map = {int(dict(zip(header, l.rstrip('\n').split('\t')))['序号']): dict(zip(header, l.rstrip('\n').split('\t'))) for l in f}
    A = {json.loads(l)['序号']: json.loads(l) for l in open(f'{BASE}/cache/passA_intro/{b}_v3merged.jsonl', encoding='utf-8')}
    Bm = {json.loads(l)['序号']: json.loads(l) for l in open(f'{BASE}/cache/passB_rules/{b}_v3merged.jsonl', encoding='utf-8')}
    for k in sorted(A):
        a, bb, s = A[k], Bm[k], S.get(k)
        # 地位词闸门：搜索终稿地位非空但无证据URL → 清空+记存疑
        status = a.get('行业地位', '') or ''
        if s and status:
            urls = [e.get('URL', '') for e in (s.get('证据') or [])]
            if not any(u.startswith('http') for u in urls):
                status = ''; gate_dropped += 1
        # 信息状态
        has_content = bool((a.get('主营业务') or '').strip())
        if s:
            info = '已核验' if s.get('地位核验') in ('确认', '修正') else ('无公开信息' if not has_content else '部分核验')
        else:
            info = '未搜索(≤100万不搜)'
        if a.get('母公司置信度') == '疑似' or (s and s.get('母公司置信度_终稿') == '疑似'):
            info = '存疑待核'
        intro = a.get('一句话简介', '')
        if not intro and not has_content:
            intro = fallback_intro(base_map[k]['行业小类'], searched=bool(s))
        parent, pnote = norm_parent((s.get('母公司_终稿') if s else '') or a.get('母公司', ''))
        parent = PARENT_ALIAS.get(parent, parent)
        parconf = (s.get('母公司置信度_终稿') if s else '') or a.get('母公司置信度', '')
        p_override = None
        _ovk = (k, a['公司名称'].strip())
        if _ovk in OVERRIDE and (OVERRIDE[_ovk].get('母公司') or '').strip():
            p_override = (OVERRIDE[_ovk]['母公司'].strip(), (OVERRIDE[_ovk].get('母公司置信度') or '').strip() or None)
        if p_override:
            parent = p_override[0]
            if p_override[1]: parconf = p_override[1]
        ev = ''
        if s and s.get('证据'):
            ev = '；'.join(f"{e.get('来源','')}:{e.get('URL','')}:{e.get('摘要','')[:30]}" for e in s['证据'][:2])
        rel = (s.get('相关度') if s else '') or a.get('相关度', '')
        # 无信息/未搜索企业：相关度不留默认R0，如实置空=未判定（用户裁定 2026-08-28）
        if rel == 'R0' and info in ('无公开信息', '未搜索(≤100万不搜)', '未核验（待补搜）'):
            rel = ''
        # QC修复1: R3/R4 不打领域标签（杂类不进标签库）
        if rel in ('R3', 'R4'):
            bb = dict(bb); bb.update({'主体系': '', '二级标签主': '', '二级标签副': '', '三级标签': '', '规则置信度': '', '标签依据': ''})
        # QC修复2: 无公开信息企业只保留登记信息标签
        if info == '无公开信息' and not (bb.get('标签依据') or '').startswith(('登记信息', '名称规则')):
            bb = dict(bb); bb.update({'主体系': '', '二级标签主': '', '二级标签副': '', '三级标签': '', '规则置信度': '', '标签依据': ''})
        # QC修复3: 地位词闸门全量执行——无搜索记录或无证据URL一律清空
        if status:
            has_ev = bool(s) and any((e.get('URL') or '').startswith('http') for e in (s.get('证据') or []))
            if not has_ev:
                status = ''; gate_dropped += 1
        # QC修复4: 最高级地位词仅弱信源 → 标存疑待核
        weak_status = bool(status) and SUPER.search(status) and WEAK_SRC.search(ev) and not STRONG_SRC.search(ev)
        remark = ((s.get('备注') if s else '') or '')
        if pnote:
            remark = (remark + '；' + pnote).strip('；')
        if weak_status:
            info = '存疑待核'
            remark = (remark + '；地位词信源偏弱待核').strip('；')
        _ovk = (k, a['公司名称'].strip())
        if _ovk in OVERRIDE:  # 人工复核修正优先（序号+公司名双键，防跨名单碰撞）
            ov = OVERRIDE[_ovk]
            bb = dict(bb)
            _flds = {'主体系': '主体系', '二级标签主': '二级标签主', '三级标签': '三级标签',
                     '标签置信度': '规则置信度', '标签依据': '标签依据'}
            for _src, _dst in _flds.items():
                if (ov.get(_src) or '').strip():   # 空字段不覆盖（只修要修的）
                    bb[_dst] = ov[_src].strip()
            if any((ov.get(_f) or '').strip() for _f in ('主体系','二级标签主','三级标签')):
                bb['标签依据'] = ov['标签依据']
        _ns, _nl1, _nl2, _nl3, _ = tag_norm(bb['主体系'], bb['二级标签主'], bb['三级标签'])
        if _nl1 is None:
            _ns, _nl1, _nl2, _nl3 = bb['主体系'], lvl1(bb['主体系'], bb['二级标签主']), bb['二级标签主'], bb['三级标签']
        if _ovk in OVERRIDE and (OVERRIDE[_ovk].get('一级') or '').strip():
            _nl1 = OVERRIDE[_ovk]['一级'].strip()
        ws.append([k, a['公司名称'], rel, (s.get('边界理由') if s else '') or '',
                   _ns, _nl1, _nl2, bb['二级标签副'], _nl3,
                   bb['规则置信度'], bb.get('标签依据', ''), norm_intro(intro), status,
                   parent, parconf, a['知名度'], (s.get('层级') if s else '未搜索'), info,
                   ev, remark, TODAY])
        n += 1
        qc[f"相关度_{rel or '空'}"] += 1
        qc[f"体系_{bb['主体系'] or '未命中'}"] += 1
        qc[f"信息状态_{info}"] += 1
        if rel in ('R3', 'R4'):
            boundary.append((k, a['公司名称'], rel, (s.get('边界理由') if s else '') or ''))
        if info == '无公开信息':
            noinfo_list.append((k, a['公司名称'], base_map[k]['注册资本万']))
        if info == '存疑待核':
            doubts.append((k, a['公司名称']))
ws.freeze_panes = 'C2'
widths = [6, 28, 6, 20, 9, 14, 17, 14, 14, 8, 10, 46, 26, 24, 9, 7, 8, 10, 50, 30, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ---------- Sheet1: 原表+标注（原表列序不动、原始顺序、右侧拼接标注列）----------
SRC = os.environ.get('MHB_SOURCE_XLSX', '')  # 源表路径，必填环境变量
swb = openpyxl.load_workbook(SRC, read_only=True)
sws = swb['P1强相关企业']
src_rows = list(sws.iter_rows(values_only=True))
NEW_COLS = ['相关度', '主体系', '一级', '二级标签主', '三级标签', '标签置信度', '一句话简介', '行业地位', '母公司', '母公司置信度', '搜索层级', '信息状态']
ws0 = wb.create_sheet('原表+标注', 0)
ws0.append(list(src_rows[0]) + NEW_COLS)
for c in ws0[1]:
    c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1F6FB2')
annot = {}
for tsv in sorted(glob.glob(f'{BASE}/full/batches/batch_f*.tsv')):
    b = tsv.split('/')[-1].replace('.tsv', '')
    with open(tsv, encoding='utf-8') as f:
        hd = f.readline().rstrip('\n').split('\t')
        hy_map = {}
        for l in f:
            d = dict(zip(hd, l.rstrip('\n').split('\t')))
            hy_map[int(d['序号'])] = d['行业小类']
    A = {json.loads(l)['序号']: json.loads(l) for l in open(f'{BASE}/cache/passA_intro/{b}_v3merged.jsonl', encoding='utf-8')}
    Bm = {json.loads(l)['序号']: json.loads(l) for l in open(f'{BASE}/cache/passB_rules/{b}_v3merged.jsonl', encoding='utf-8')}
    for k in sorted(A):
        a, bb, s = A[k], Bm[k], S.get(k)
        parent, _ = norm_parent((s.get('母公司_终稿') if s else '') or a.get('母公司', ''))
        parent = PARENT_ALIAS.get(parent, parent)
        _ovk = (k, a['公司名称'].strip())
        if _ovk in OVERRIDE and (OVERRIDE[_ovk].get('母公司') or '').strip():
            parent = OVERRIDE[_ovk]['母公司'].strip()
        has_content = bool((a.get('主营业务') or '').strip())
        if s:
            info = '已核验' if s.get('地位核验') in ('确认', '修正') else ('无公开信息' if not has_content else '部分核验')
        else:
            info = '未搜索(≤100万不搜)'
        if a.get('母公司置信度') == '疑似' or (s and s.get('母公司置信度_终稿') == '疑似'):
            info = '存疑待核'
        rel = (s.get('相关度') if s else '') or a.get('相关度', '')
        # 无信息/未搜索企业：相关度不留默认R0，如实置空=未判定（用户裁定 2026-08-28）
        if rel == 'R0' and info in ('无公开信息', '未搜索(≤100万不搜)', '未核验（待补搜）'):
            rel = ''
        if rel in ('R3', 'R4'):
            bb = dict(bb); bb.update({'主体系': '', '二级标签主': '', '三级标签': '', '规则置信度': ''})
        if info == '无公开信息' and not (bb.get('标签依据') or '').startswith(('登记信息', '名称规则')):
            bb = dict(bb); bb.update({'主体系': '', '二级标签主': '', '三级标签': '', '规则置信度': ''})
        _ovk = (k, a['公司名称'].strip())
        if _ovk in OVERRIDE:  # 人工复核修正优先（序号+公司名双键，防跨名单碰撞）
            ov = OVERRIDE[_ovk]
            bb = dict(bb)
            _flds = {'主体系': '主体系', '二级标签主': '二级标签主', '三级标签': '三级标签',
                     '标签置信度': '规则置信度', '标签依据': '标签依据'}
            for _src, _dst in _flds.items():
                if (ov.get(_src) or '').strip():   # 空字段不覆盖（只修要修的）
                    bb[_dst] = ov[_src].strip()
            if any((ov.get(_f) or '').strip() for _f in ('主体系','二级标签主','三级标签')):
                bb['标签依据'] = ov['标签依据']
        status = a.get('行业地位', '') or ''
        if status:
            has_ev = bool(s) and any((e.get('URL') or '').startswith('http') for e in (s.get('证据') or []))
            if not has_ev: status = ''
        intro = a.get('一句话简介', '')
        if not intro and not has_content:
            intro = fallback_intro(hy_map.get(k, ''), searched=bool(s))
        _ns, _nl1, _nl2, _nl3, _ = tag_norm(bb['主体系'], bb['二级标签主'], bb['三级标签'])
        if _nl1 is None:
            _ns, _nl1, _nl2, _nl3 = bb['主体系'], lvl1(bb['主体系'], bb['二级标签主']), bb['二级标签主'], bb['三级标签']
        annot[k] = [rel, _ns, _nl1, _nl2, _nl3,
                    bb['规则置信度'], norm_intro(intro), status, parent,
                    (s.get('母公司置信度_终稿') if s else '') or a.get('母公司置信度', ''),
                    (s.get('层级') if s else '未搜索'), info]
n_orig = 0
for row in src_rows[1:]:
    sid = row[0]
    ann = annot.get(sid)
    if ann is None:
        ann = ['', '', '', '', '', '', '', '', '', '', '', '非存续（未标注）']
    ws0.append(list(row) + ann)
    n_orig += 1
ws0.freeze_panes = 'C2'
ws0.column_dimensions['B'].width = 30
print(f'原表+标注 sheet: {n_orig} 行（含非存续 5 家）')

FINAL_XLSX = os.environ.get('MHB_FINAL_XLSX', f'{BASE}/full/终稿_企业标注.xlsx')
wb.save(FINAL_XLSX)

print(f'终稿行数: {n}, 地位词闸门清空: {gate_dropped}')
for k, v in sorted(qc.items()):
    print(f'  {k}: {v}')
# 质检清单落盘
with open(f'{BASE}/full/质检_边界企业清单.tsv', 'w', encoding='utf-8') as f:
    f.write('序号\t公司名称\t相关度\t边界理由\n')
    for x in boundary: f.write('\t'.join(map(str, x)) + '\n')
with open(f'{BASE}/full/质检_无公开信息清单.tsv', 'w', encoding='utf-8') as f:
    f.write('序号\t公司名称\t注册资本万\n')
    for x in sorted(noinfo_list, key=lambda x: int(x[2])): f.write('\t'.join(map(str, x)) + '\n')
with open(f'{BASE}/full/质检_存疑待核清单.tsv', 'w', encoding='utf-8') as f:
    f.write('序号\t公司名称\n')
    for x in doubts: f.write('\t'.join(map(str, x)) + '\n')
print(f'\n质检清单: 边界 {len(boundary)}, 无公开信息 {len(noinfo_list)}, 存疑待核 {len(doubts)}')
