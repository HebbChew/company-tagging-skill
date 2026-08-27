#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""阶段Ⅰ收尾：全量分层 + 标签版Excel。用法: python3 merge_stage1.py"""
import json, csv, glob, os, collections, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import os
BASE = os.environ.get('MHB_TAG_BASE', os.path.dirname(os.path.abspath(__file__)))
os.makedirs(f'{BASE}/full/分层', exist_ok=True)

# ---------- 1) 分层 ----------
def kc90(s):
    s = str(s or '').strip()
    return s.isdigit() and int(s) >= 90
tier_stat = collections.Counter()
for tsv in sorted(glob.glob(f'{BASE}/full/batches/batch_f*.tsv')):
    b = tsv.split('/')[-1].replace('.tsv', '')
    flags = {}
    with open(tsv, encoding='utf-8') as f:
        header = f.readline().rstrip('\n').split('\t')
        for line in f:
            p = line.rstrip('\n').split('\t')
            d = dict(zip(header, p))
            flags[int(d['序号'])] = d
    out = []
    for l in open(f'{BASE}/cache/passA_intro/{b}_v3.jsonl', encoding='utf-8'):
        a = json.loads(l)
        fl = flags[a['序号']]
        hard = fl['专精特新'] == '是' or float(fl['营收亿'] or 0) >= 1 or kc90(fl['科创分']) or float(fl['注册资本万'] or 0) >= 10000
        if a['知名度'] in ('普通', '未知') and hard:
            tier = 'S2'
        elif a['知名度'] == '知名' or a.get('母公司置信度') == '疑似' or fl['规上'] == '是':
            tier = 'S1'
        else:
            tier = 'S0'
        tier_stat[tier] += 1
        out.append({'序号': a['序号'], '公司名称': a['公司名称'], '层级': tier, '知名度': a['知名度'],
                    '硬信号': ('专精特新' if fl['专精特新'] == '是' else '') + ('规上' if fl['规上'] == '是' else '') +
                              (f"营收{fl['营收亿']}亿" if float(fl['营收亿'] or 0) >= 1 else '') +
                              (f"科创{fl['科创分']}" if kc90(fl['科创分']) else '') +
                              (f"注册{float(fl['注册资本万'])/10000:.1f}亿" if float(fl['注册资本万'] or 0) >= 10000 else ''),
                    '行业地位': a.get('行业地位', ''), '母公司': a.get('母公司', ''), '母公司置信度': a.get('母公司置信度', '')})
    with open(f'{BASE}/full/分层/{b}_分层.jsonl', 'w', encoding='utf-8') as f:
        for o in out:
            f.write(json.dumps(o, ensure_ascii=False) + '\n')
print('分层分布:', dict(tier_stat))

# ---------- 2) 一级映射 ----------
m12, sys2l1 = {}, {}
with open(f'{BASE}/技术标签词表.csv', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        m12[(r['体系'], r['二级'])] = r['一级']
        sys2l1.setdefault(r['体系'], r['一级'])
SYNBIO_L2 = {'材料应用', '食品应用', '农牧应用', '其他应用'}
def lvl1(s, l2):
    if s == '周边配套': return '周边配套'
    if l2 in ('医药供应链上游配套', '医药流通与供应链'): return '医药供应链与流通（补充）'
    if l2 in SYNBIO_L2: return '合成生物学'
    return m12.get((s, l2), sys2l1.get(s, ''))

# ---------- 3) 合并标签版 ----------
wb = openpyxl.Workbook(); ws = wb.active; ws.title = '标签版(阶段Ⅰ)'
HDR = ['序号', '公司名称', '相关度', '主体系', '一级', '二级标签主', '二级标签副', '三级标签', '标签置信度', '标签依据',
       '一句话简介(初判)', '行业地位(初判)', '母公司(初判)', '母公司置信度', '知名度', '搜索层级', '依据']
ws.append(HDR)
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1F6FB2')
    c.alignment = Alignment(horizontal='center', vertical='center')
n = 0
for tsv in sorted(glob.glob(f'{BASE}/full/batches/batch_f*.tsv')):
    b = tsv.split('/')[-1].replace('.tsv', '')
    A = {json.loads(l)['序号']: json.loads(l) for l in open(f'{BASE}/cache/passA_intro/{b}_v3.jsonl', encoding='utf-8')}
    B = {json.loads(l)['序号']: json.loads(l) for l in open(f'{BASE}/cache/passB_rules/{b}_v3.jsonl', encoding='utf-8')}
    T = {json.loads(l)['序号']: json.loads(l) for l in open(f'{BASE}/full/分层/{b}_分层.jsonl', encoding='utf-8')}
    for k in sorted(A):
        a, bb, t = A[k], B[k], T[k]
        ws.append([k, a['公司名称'], a.get('相关度', ''), bb['主体系'], lvl1(bb['主体系'], bb['二级标签主']),
                   bb['二级标签主'], bb['二级标签副'], bb['三级标签'], bb['规则置信度'], bb.get('标签依据', ''),
                   a.get('一句话简介', ''), a.get('行业地位', ''), a.get('母公司', ''), a.get('母公司置信度', ''),
                   a['知名度'], t['层级'], a.get('依据', '')])
        n += 1
ws.freeze_panes = 'C2'
for i, w in enumerate([6, 30, 7, 9, 14, 17, 15, 15, 8, 46, 24, 24, 9, 7, 7, 22], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
wb.save(f'{BASE}/full/标签版_阶段1.xlsx')
print('标签版行数:', n)
