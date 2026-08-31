#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""对抗审查器：终稿交付前的自动红队检查（四层复核体系的 L3 自动化部分）。
检查"展示字段与证据的一致性"，规则来自闵行区项目实战漏洞复盘（2026-08-31 终版）。

用法: MHB_TAG_BASE=<目录> MHB_FINAL_XLSX=<终稿路径> python3 qc_adversarial.py
退出码: 0=全部通过或有误报级项; 1=存在必修违规
"""
import os, re, sys
from collections import Counter
import openpyxl

BASE = os.environ.get('MHB_TAG_BASE', os.path.dirname(os.path.abspath(__file__)))
FINAL = os.environ.get('MHB_FINAL_XLSX', f'{BASE}/full/终稿_企业标注.xlsx')

PLACEHOLDER = re.compile(r'^（登记行业')
NOINFO_WORD = re.compile(r'暂无|未见|未检得')
BIZ_EV = re.compile(r'中标|官网|产品|获批|融资|专利|生产|研发|服务|平台|客户|龙头|首个|门诊|诊疗')
# 登记直标豁免（E0 级：机构类型/名称规则/母系继承，见 SKILL.md §3.6）
REG_EXEMPT = ('机构类型', '名称规则', '母系继承')
VALID_L1 = {'药物', '医疗器械', '医药研发服务', '医药供应链与流通（补充）', '合成生物与生物制造',
            '生物技术与上游试剂耗材设备', '数字医疗与AI', '周边配套', '医疗服务（补充）'}

def run(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['终稿'] if '终稿' in wb.sheetnames else wb[wb.sheetnames[-1]]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0]) if h}
    def g(r, k):
        v = r[idx[k]] if k in idx else None
        return str(v) if v is not None else ''
    def ev_sub(r):
        ev = g(r, '证据摘要')
        return ev not in ('', 'None') and not NOINFO_WORD.search(ev)

    R = Counter(); ex = {}
    def rep(rule, r):
        R[rule] += 1
        if len(ex.setdefault(rule, [])) < 5:
            ex[rule].append(g(r, '公司名称'))

    for r in rows[1:]:
        intro = g(r, '一句话简介')
        zero_ev = bool(PLACEHOLDER.search(intro)) and not ev_sub(r)
        # A 占位简介但有证据（简介盖证据 bug 的回归测试）
        if PLACEHOLDER.search(intro) and ev_sub(r):
            rep('A 占位简介盖住证据', r)
        # B 零证据但有标签/相关度
        if zero_ev and g(r, '一级') and not g(r, '标签依据').startswith(REG_EXEMPT):
            rep('B 零证据有标签', r)
        if zero_ev and g(r, '相关度'):
            rep('B2 零证据有相关度', r)
        # C 地位词必须有 URL
        if g(r, '行业地位') and 'http' not in g(r, '证据摘要'):
            rep('C 地位词无URL', r)
        # D 有证据不得标无公开信息（剥离URL后，摘要文本含业务实质词才算冲突）
        if g(r, '信息状态') == '无公开信息' and BIZ_EV.search(re.sub(r'https?://\S+', '', g(r, '证据摘要'))):
            rep('D 有业务证据却标无公开信息', r)
        # E R3/R4 必须有边界理由；未定级行的理由不得含判级断言
        if g(r, '相关度') in ('R3', 'R4') and not g(r, '边界理由'):
            rep('E R3/R4无边界理由', r)
        if not g(r, '相关度') and g(r, '边界理由'):
            rb = g(r, '边界理由')
            if not re.search(r'未定级|未达判级|待核|疑似|暂无|未见|无出界|无业务信息|宁留勿冤', rb):
                rep('E2 未定级但理由含判级断言', r)
        # F 结构：一级空但二级/置信度非空
        if not g(r, '一级') and (g(r, '二级标签主') or g(r, '标签置信度')):
            rep('F 一级空但有二级/置信度', r)
        # G 不设二级的一级带了二级
        if g(r, '一级') in ('医疗服务（补充）', '周边配套') and g(r, '二级标签主'):
            rep('G 不设二级的一级有二级', r)
        # H 母公司确定但无证据
        if g(r, '母公司置信度') == '确定' and not ev_sub(r):
            rep('H 母公司确定但无证据', r)
        # I 已核实但证据空
        if g(r, '信息状态') == '已核实' and not ev_sub(r):
            rep('I 已核实但证据空', r)
        # J 层级枚举封闭
        if g(r, '搜索层级') and g(r, '搜索层级') not in ('深挖', '核验', '打底', '追挖', '未搜索'):
            rep('J 搜索层级越界', r)
        # L 一级封闭枚举
        if g(r, '一级') and g(r, '一级') not in VALID_L1:
            rep('L 一级越界白名单', r)
        # K 未搜索与状态一致
        if g(r, '信息状态') == '未搜索(≤100万不搜)' and g(r, '搜索层级') not in ('未搜索', ''):
            rep('K 未搜索状态但层级非未搜索', r)
    return R, ex

R, ex = run(FINAL)
print('== 对抗审查 ==')
fail = 0
for k, v in R.most_common():
    print(f'[{k}] {v}  例: {ex[k]}')
    fail += v
print('全部通过 ✓' if fail == 0 else f'\n必修违规合计: {fail}')
sys.exit(1 if fail else 0)
